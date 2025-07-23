from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import openpyxl
import os
import tempfile
from werkzeug.utils import secure_filename
import shutil
import re

mercos_v3_bp = Blueprint("mercos_v3", __name__)

# Carregar a base de dados de produtos uma vez quando o módulo é importado
BASE_PRODUTOS_PATH = "RELACAODEPRODUTOSMAG.xlsx"
TEMPLATE_PEQPAPER_PATH = "16249.xlsx"

def carregar_base_produtos():
    """Carrega a base de dados de produtos e cria o mapeamento EAN -> Código Ommie"""
    try:
        df_produtos = pd.read_excel(BASE_PRODUTOS_PATH, header=1)
        
        # Limpar dados nulos e criar mapeamento
        df_produtos = df_produtos.dropna(subset=["Código EAN (GTIN)", "Código do Produto"])
        
        # Converter códigos EAN para string para evitar problemas de tipo
        df_produtos["Código EAN (GTIN)"] = df_produtos["Código EAN (GTIN)"].astype(str)
        df_produtos["Código do Produto"] = df_produtos["Código do Produto"].astype(str)
        
        # Criar dicionário de mapeamento: EAN -> Código Ommie
        mapeamento = dict(zip(df_produtos["Código EAN (GTIN)"], df_produtos["Código do Produto"]))
        
        return mapeamento, len(df_produtos)
    except Exception as e:
        print(f"Erro ao carregar base de produtos: {e}")
        return {}, 0

# Carregar a base de produtos na inicialização
MAPEAMENTO_EAN_OMMIE, TOTAL_PRODUTOS = carregar_base_produtos()

@mercos_v3_bp.route("/automatizar-planilha-peqpaper", methods=["POST"])
def automatizar_mercos_planilha_peqpaper():
    try:
        # Verificar se o arquivo foi enviado
        if "orcamento" not in request.files:
            return jsonify({"error": "Arquivo de orçamento é obrigatório"}), 400
        
        orcamento_file = request.files["orcamento"]
        
        if orcamento_file.filename == "":
            return jsonify({"error": "Nenhum arquivo foi selecionado"}), 400
        
        # Verificar se a base de produtos foi carregada
        if not MAPEAMENTO_EAN_OMMIE:
            return jsonify({"error": "Base de dados de produtos não encontrada"}), 500
        
        # Verificar se o template Planilha peqpaper existe
        if not os.path.exists(TEMPLATE_PEQPAPER_PATH):
            return jsonify({"error": "Template da Planilha peqpaper não encontrado"}), 500
        
        # Criar diretório temporário para processar os arquivos
        with tempfile.TemporaryDirectory() as temp_dir:
            # Salvar arquivo temporariamente
            orcamento_path = os.path.join(temp_dir, secure_filename(orcamento_file.filename))
            template_path = os.path.join(temp_dir, "planilha_peqpaper_template.xlsx")
            saida_path = os.path.join(temp_dir, "planilha_peqpaper_preenchida.xlsx")
            
            orcamento_file.save(orcamento_path)
            
            # Copiar o template Planilha peqpaper para o diretório temporário
            shutil.copy2(TEMPLATE_PEQPAPER_PATH, template_path)
            
            # Processar a planilha de orçamento
            df_orcamento = pd.read_excel(orcamento_path, header=None, skiprows=4)
            df_orcamento.rename(columns={1: "CODIGO_BARRAS", 3: "QUANTIDADE", 7: "PRECO"}, inplace=True)
            df_orcamento = df_orcamento[["CODIGO_BARRAS", "QUANTIDADE", "PRECO"]].dropna(subset=["CODIGO_BARRAS"])
            
            # Converter códigos de barra corretamente
            df_orcamento["CODIGO_BARRAS"] = df_orcamento["CODIGO_BARRAS"].astype("int64").astype(str)
            
            # Mapear códigos de barra para códigos Ommie
            df_orcamento["CODIGO_OMMIE"] = df_orcamento["CODIGO_BARRAS"].map(MAPEAMENTO_EAN_OMMIE)
            
            # Filtrar apenas produtos encontrados
            df_encontrados = df_orcamento.dropna(subset=["CODIGO_OMMIE"])
            
            # Carregar o template da Planilha peqpaper
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # --- Extrair CNPJ da planilha de orçamento e preencher D7 ---
            df_orcamento_full = pd.read_excel(orcamento_path, header=None) # Ler a planilha completa para buscar o CNPJ
            cnpj_encontrado = None
            cnpj_pattern = r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}"
            
            for r_idx in range(df_orcamento_full.shape[0]):
                for c_idx in range(df_orcamento_full.shape[1]):
                    cell_value = str(df_orcamento_full.iloc[r_idx, c_idx])
                    match = re.search(cnpj_pattern, cell_value)
                    if match:
                        cnpj_encontrado = match.group(0)
                        break
                if cnpj_encontrado:
                    break
            
            if cnpj_encontrado:
                ws["D7"] = cnpj_encontrado
            else:
                print("Aviso: CNPJ não encontrado na planilha de orçamento.")
            # ---------------------------------------------------------
            
            # Preencher os dados nas posições corretas
            # D22 para baixo: Códigos Ommie
            # F22 para baixo: Quantidades
            # G22 para baixo: Preços
            
            linha_inicial = 22
            for idx, row in df_encontrados.iterrows():
                linha_atual = linha_inicial + len(df_encontrados[:idx])
                
                # Coluna D: Código Ommie
                ws[f"D{linha_atual}"] = row["CODIGO_OMMIE"]
                
                # Coluna F: Quantidade
                ws[f"F{linha_atual}"] = row["QUANTIDADE"]
                
                # Coluna G: Preço
                ws[f"G{linha_atual}"] = row["PRECO"]
            
            # Salvar a planilha preenchida
            wb.save(saida_path)
            
            # Preparar informações de retorno
            total_produtos_orcamento = len(df_orcamento)
            produtos_encontrados = len(df_encontrados)
            produtos_nao_encontrados = total_produtos_orcamento - produtos_encontrados
            
            info_processamento = {
                "total_produtos_orcamento": total_produtos_orcamento,
                "produtos_encontrados": produtos_encontrados,
                "produtos_nao_encontrados": produtos_nao_encontrados,
                "taxa_sucesso": f"{(produtos_encontrados/total_produtos_orcamento)*100:.1f}%",
                "base_produtos_total": TOTAL_PRODUTOS,
                "linhas_preenchidas": f"D{linha_inicial} até D{linha_inicial + produtos_encontrados - 1}"
            }
            
            # Se houver produtos não encontrados, incluir a lista
            if produtos_nao_encontrados > 0:
                produtos_nao_encontrados_df = df_orcamento[df_orcamento["CODIGO_OMMIE"].isna()]
                codigos_nao_encontrados = produtos_nao_encontrados_df["CODIGO_BARRAS"].tolist()
                info_processamento["codigos_nao_encontrados"] = codigos_nao_encontrados[:10]
            
            # Retornar o arquivo para download
            response = send_file(saida_path, as_attachment=True, download_name="planilha_peqpaper_preenchida.xlsx")
            response.headers["X-Processing-Info"] = str(info_processamento)
            
            return response
            
    except Exception as e:
        return jsonify({"error": f"Erro ao processar arquivo: {str(e)}"}), 500

@mercos_v3_bp.route("/status-template", methods=["GET"])
def status_template():
    """Retorna informações sobre o template e base de dados"""
    template_existe = os.path.exists(TEMPLATE_PEQPAPER_PATH)
    base_carregada = len(MAPEAMENTO_EAN_OMMIE) > 0
    
    return jsonify({
        "status": "Sistema pronto" if template_existe and base_carregada else "Sistema com problemas",
        "template_peqpaper_existe": template_existe,
        "base_produtos_carregada": base_carregada,
        "total_produtos": TOTAL_PRODUTOS
    })


